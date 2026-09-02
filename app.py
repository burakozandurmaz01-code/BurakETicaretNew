from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity, verify_jwt_in_request
from werkzeug.exceptions import HTTPException
import sqlite3
import bcrypt
import os
import uuid
from datetime import datetime, timedelta
from functools import wraps
import hashlib
import io
import json
import re
import traceback
from html import escape
from openpyxl import Workbook

try:
    import psycopg2
    _IntegrityError = (sqlite3.IntegrityError, psycopg2.IntegrityError)
except ImportError:
    _IntegrityError = (sqlite3.IntegrityError,)
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

app = Flask(__name__)
CORS(app)

# Configuration
app.config['JWT_SECRET_KEY'] = os.environ.get('JWT_SECRET_KEY', 'burak-eticaret-secret-key-2024')
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=24)
jwt = JWTManager(app)

# Database - SQLite locally, PostgreSQL on Render (via DATABASE_URL)
DATABASE_URL = os.environ.get('DATABASE_URL')
DATABASE_PATH = os.path.join(os.path.dirname(__file__), 'database', 'burak_eticaret.db')

if not os.path.exists('database'):
    os.makedirs('database')

class CursorWrapper:
    def __init__(self, cursor, db_type):
        self._cursor = cursor
        self._db_type = db_type

    def execute(self, query, params=None):
        if self._db_type == 'postgres':
            query = query.replace('?', '%s')
            if params is not None:
                params = tuple(int(p) if isinstance(p, bool) else p for p in params)
        if params is None:
            return self._cursor.execute(query)
        return self._cursor.execute(query, params)

    def fetchone(self):
        return self._cursor.fetchone()

    def fetchall(self):
        return self._cursor.fetchall()

    def close(self):
        return self._cursor.close()

    def __getattr__(self, name):
        return getattr(self._cursor, name)

class ConnectionWrapper:
    def __init__(self, conn, db_type):
        self._conn = conn
        self._db_type = db_type

    def cursor(self):
        return CursorWrapper(self._conn.cursor(), self._db_type)

    def commit(self):
        return self._conn.commit()

    def close(self):
        return self._conn.close()

    def __getattr__(self, name):
        return getattr(self._conn, name)

def get_db():
    if DATABASE_URL:
        import psycopg2
        import psycopg2.extras
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
        conn.autocommit = False
        return ConnectionWrapper(conn, 'postgres')
    else:
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        conn.execute('PRAGMA foreign_keys = ON')
        return ConnectionWrapper(conn, 'sqlite')

# ---- Helpers and utilities ----

def _get_connection_type(conn):
    return getattr(conn, '_db_type', 'sqlite')

def column_exists(conn, table, column):
    """Check whether a column exists in the connected database."""
    cursor = conn.cursor()
    db_type = _get_connection_type(conn)
    if db_type == 'postgres':
        cursor.execute('''
            SELECT 1 FROM information_schema.columns
            WHERE table_name = %s AND column_name = %s
        ''', (table, column))
    else:
        cursor.execute(f'PRAGMA table_info({table})')
        for row in cursor.fetchall():
            if row['name'] == column:
                return True
        return False
    return cursor.fetchone() is not None

def add_column_if_missing(conn, table, column, definition):
    if not column_exists(conn, table, column):
        cursor = conn.cursor()
        try:
            cursor.execute(f'ALTER TABLE {table} ADD COLUMN {column} {definition}')
        except Exception:
            pass

def migrate_db():
    """Apply non-destructive schema migrations for new features."""
    conn = get_db()
    cursor = conn.cursor()

    # SKU registry ensures deleted SKUs are never reused
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sku_registry (
            sku TEXT PRIMARY KEY,
            entity_type TEXT,
            entity_id TEXT,
            deleted_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Products cost/tracking columns
    add_column_if_missing(conn, 'products', 'low_stock_threshold', 'INTEGER DEFAULT 10')
    add_column_if_missing(conn, 'products', 'packaging_cost', 'REAL DEFAULT 0')
    add_column_if_missing(conn, 'products', 'commission', 'REAL DEFAULT 0')
    add_column_if_missing(conn, 'products', 'other_costs', 'REAL DEFAULT 0')
    add_column_if_missing(conn, 'products', 'deleted_at', 'TIMESTAMP')

    # Soft delete columns
    add_column_if_missing(conn, 'customers', 'deleted_at', 'TIMESTAMP')
    add_column_if_missing(conn, 'orders', 'deleted_at', 'TIMESTAMP')
    add_column_if_missing(conn, 'categories', 'deleted_at', 'TIMESTAMP')
    add_column_if_missing(conn, 'categories', 'updated_at', 'TIMESTAMP')

    # Order shipping tracking columns
    add_column_if_missing(conn, 'orders', 'tracking_number', 'TEXT')
    add_column_if_missing(conn, 'orders', 'shipping_company', 'TEXT')
    add_column_if_missing(conn, 'orders', 'shipping_status', 'TEXT DEFAULT \'pending\'')

    # Stock movements created_by already in schema; ensure deleted_at not needed here
    conn.commit()
    conn.close()

_TR_TRANSLATION = str.maketrans('çğıöşüÇĞİÖŞÜ', 'cgiosuCGIOSU')

def normalize_for_sku(text):
    if not text:
        return ''
    text = str(text).translate(_TR_TRANSLATION)
    text = re.sub(r'[^A-Za-z0-9]+', '-', text).strip('-').upper()
    return text[:25]

def sku_exists(sku):
    """Return True if SKU is already reserved anywhere in the system."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT 1 FROM (
            SELECT sku FROM sku_registry WHERE sku = ?
            UNION
            SELECT sku FROM products WHERE sku = ? AND deleted_at IS NULL
            UNION
            SELECT sku FROM product_variations WHERE sku = ?
        ) AS reserved
    ''', (sku, sku, sku))
    exists = cursor.fetchone() is not None
    conn.close()
    return exists

def generate_sku(name, category=None, variation_type=None):
    """Generate a unique, human-readable SKU."""
    parts = [normalize_for_sku(p) for p in (category, name, variation_type) if p]
    prefix = '-'.join(parts) or 'URUN'
    prefix = prefix[:40]

    candidate = prefix
    counter = 0
    while sku_exists(candidate):
        counter += 1
        candidate = f"{prefix}-{counter:03d}"
    return candidate

def register_sku(sku, entity_type, entity_id, deleted_at=None):
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO sku_registry (sku, entity_type, entity_id, deleted_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT (sku) DO UPDATE SET
                entity_type = EXCLUDED.entity_type,
                entity_id = EXCLUDED.entity_id,
                deleted_at = EXCLUDED.deleted_at
        ''', (sku, entity_type, entity_id, deleted_at))
        conn.commit()
    except _IntegrityError:
        pass
    conn.close()

def get_current_user():
    try:
        from flask_jwt_extended import get_jwt
        return get_jwt()
    except Exception:
        return {}

def get_current_user_id():
    try:
        return get_jwt_identity() or request.headers.get('X-User-ID', 'admin')
    except Exception:
        return request.headers.get('X-User-ID', 'admin')

def get_current_user_role():
    try:
        from flask_jwt_extended import get_jwt
        return get_jwt().get('role', 'admin')
    except Exception:
        return 'admin'

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        role = get_current_user_role()
        if role != 'admin':
            return jsonify({'error': 'Bu işlem için yönetici yetkisi gerekli'}), 403
        return f(*args, **kwargs)
    return decorated_function

def log_activity(action, entity_type=None, entity_id=None, details=None):
    """Record an activity log entry for the current user."""
    user_id = get_current_user_id()
    if not user_id:
        return
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO activity_logs (id, user_id, action, entity_type, entity_id, details, ip_address, user_agent)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (str(uuid.uuid4()), user_id, action, entity_type, entity_id,
              json.dumps(details, ensure_ascii=False, default=str) if details else None,
              request.remote_addr, request.headers.get('User-Agent')))
        conn.commit()
    except Exception:
        pass
    conn.close()

def _insert_stock_movement(cursor, product_id, quantity, movement_type, reference_id=None, reference_type=None, notes=None):
    """Insert a stock movement row using an existing cursor."""
    user_id = get_current_user_id()
    cursor.execute('''
        INSERT INTO stock_movements (id, product_id, movement_type, quantity, reference_id, reference_type, notes, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (str(uuid.uuid4()), product_id, movement_type, quantity, reference_id, reference_type, notes, user_id))

def record_stock_movement(product_id, quantity, movement_type, reference_id=None, reference_type=None, notes=None, adjust=True):
    """Record a single stock movement and optionally update product stock_quantity."""
    conn = get_db()
    cursor = conn.cursor()
    if adjust:
        cursor.execute('UPDATE products SET stock_quantity = stock_quantity + ? WHERE id = ?', (quantity, product_id))
    _insert_stock_movement(cursor, product_id, quantity, movement_type, reference_id, reference_type, notes)
    conn.commit()
    conn.close()

def _as_dict(row):
    return row if hasattr(row, 'get') else dict(row)

def get_product_costs(product):
    """Return per-unit costs for a product, including default zeros."""
    product = _as_dict(product)
    return {
        'cost_price': float(product.get('cost_price') or 0),
        'packaging_cost': float(product.get('packaging_cost') or 0),
        'commission': float(product.get('commission') or 0),
        'other_costs': float(product.get('other_costs') or 0),
    }

def calculate_product_profit(product):
    product = _as_dict(product)
    price = float(product.get('price') or 0)
    costs = get_product_costs(product)
    total_cost = sum(costs.values())
    return round(price - total_cost, 2)

def calculate_order_profit(order, items):
    """Calculate net profit for an order based on product costs."""
    order = _as_dict(order)
    conn = get_db()
    cursor = conn.cursor()
    total_cost = 0
    for it in items:
        it = _as_dict(it)
        cursor.execute('SELECT price, cost_price, packaging_cost, commission, other_costs FROM products WHERE id = ?', (it.get('product_id'),))
        product = cursor.fetchone()
        if product:
            costs = get_product_costs(product)
            total_cost += (costs['cost_price'] + costs['packaging_cost'] + costs['commission'] + costs['other_costs']) * it.get('quantity', 0)
    conn.close()
    revenue = float(order.get('total') or 0)
    shipping = float(order.get('shipping_cost') or 0)
    discount = float(order.get('discount') or 0)
    return round(revenue - total_cost - shipping - discount, 2)

def _serialize_row(row):
    d = dict(row)
    for k, v in d.items():
        if isinstance(v, datetime):
            d[k] = v.isoformat()
    return d

def _upsert_table(cursor, table, rows, pk='id'):
    if not rows:
        return
    columns = [c for c in rows[0].keys()]
    cols_str = ', '.join(columns)
    placeholders = ', '.join('?' * len(columns))
    update_cols = [c for c in columns if c != pk]
    if update_cols:
        update_set = ', '.join(f'{c} = EXCLUDED.{c}' for c in update_cols)
        query = f'INSERT INTO {table} ({cols_str}) VALUES ({placeholders}) ON CONFLICT ({pk}) DO UPDATE SET {update_set}'
    else:
        query = f'INSERT INTO {table} ({cols_str}) VALUES ({placeholders}) ON CONFLICT ({pk}) DO NOTHING'
    for row in rows:
        cursor.execute(query, tuple(row[c] for c in columns))

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads', 'products')
BACKUP_DIR = os.path.join(os.path.dirname(__file__), 'backups')

# Ensure directories exist
os.makedirs(os.path.dirname(DATABASE_PATH), exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

# Register PDF fonts with Turkish character support
FONT_DIR = os.path.join(os.path.dirname(__file__), 'fonts')
try:
    pdfmetrics.registerFont(TTFont('DejaVuSans', os.path.join(FONT_DIR, 'DejaVuSans.ttf')))
    pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', os.path.join(FONT_DIR, 'DejaVuSans-Bold.ttf')))
except Exception:
    pass

def init_db():
    conn = get_db()
    cursor = conn.cursor()
    
    # Users table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            email TEXT UNIQUE,
            full_name TEXT,
            role TEXT DEFAULT 'user',
            theme TEXT DEFAULT 'light',
            language TEXT DEFAULT 'tr',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Categories table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS categories (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            description TEXT,
            parent_id TEXT,
            image_url TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (parent_id) REFERENCES categories(id)
        )
    ''')
    
    # Products table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS products (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            description TEXT,
            price REAL NOT NULL,
            cost_price REAL,
            stock_quantity INTEGER DEFAULT 0,
            category_id TEXT,
            sku TEXT UNIQUE,
            barcode TEXT,
            image_url TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (category_id) REFERENCES categories(id)
        )
    ''')
    
    # Customers table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS customers (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            email TEXT,
            phone TEXT,
            address TEXT,
            city TEXT,
            tax_number TEXT,
            tax_office TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Orders table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS orders (
            id TEXT PRIMARY KEY,
            customer_id TEXT,
            order_number TEXT UNIQUE NOT NULL,
            status TEXT DEFAULT 'pending',
            subtotal REAL NOT NULL,
            tax REAL DEFAULT 0,
            shipping_cost REAL DEFAULT 0,
            discount REAL DEFAULT 0,
            total REAL NOT NULL,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (customer_id) REFERENCES customers(id)
        )
    ''')
    
    # Order items table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS order_items (
            id TEXT PRIMARY KEY,
            order_id TEXT NOT NULL,
            product_id TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            unit_price REAL NOT NULL,
            total_price REAL NOT NULL,
            FOREIGN KEY (order_id) REFERENCES orders(id),
            FOREIGN KEY (product_id) REFERENCES products(id)
        )
    ''')
    
    # Settings table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Product variations table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS product_variations (
            id TEXT PRIMARY KEY,
            product_id TEXT NOT NULL,
            phone_model TEXT,
            color TEXT,
            protection_type TEXT,
            price REAL NOT NULL,
            stock_quantity INTEGER DEFAULT 0,
            sku TEXT,
            barcode TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (product_id) REFERENCES products(id)
        )
    ''')
    
    # Stock movements table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS stock_movements (
            id TEXT PRIMARY KEY,
            product_id TEXT NOT NULL,
            movement_type TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            reference_id TEXT,
            reference_type TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            created_by TEXT,
            FOREIGN KEY (product_id) REFERENCES products(id),
            FOREIGN KEY (created_by) REFERENCES users(id)
        )
    ''')
    
    # Suppliers table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS suppliers (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            contact_person TEXT,
            email TEXT,
            phone TEXT,
            address TEXT,
            city TEXT,
            tax_number TEXT,
            tax_office TEXT,
            payment_terms TEXT,
            notes TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Returns/Refunds table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS returns (
            id TEXT PRIMARY KEY,
            order_id TEXT NOT NULL,
            customer_id TEXT,
            return_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            return_type TEXT NOT NULL,
            reason TEXT,
            status TEXT DEFAULT 'pending',
            refund_amount REAL,
            refund_method TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (order_id) REFERENCES orders(id),
            FOREIGN KEY (customer_id) REFERENCES customers(id)
        )
    ''')
    
    # Return items table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS return_items (
            id TEXT PRIMARY KEY,
            return_id TEXT NOT NULL,
            product_id TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            reason TEXT,
            condition TEXT,
            FOREIGN KEY (return_id) REFERENCES returns(id),
            FOREIGN KEY (product_id) REFERENCES products(id)
        )
    ''')
    
    # Coupons table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS coupons (
            id TEXT PRIMARY KEY,
            code TEXT UNIQUE NOT NULL,
            discount_type TEXT NOT NULL,
            discount_value REAL NOT NULL,
            minimum_purchase REAL DEFAULT 0,
            max_discount REAL,
            usage_limit INTEGER,
            used_count INTEGER DEFAULT 0,
            valid_from TIMESTAMP,
            valid_until TIMESTAMP,
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Finance transactions table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS finance_transactions (
            id TEXT PRIMARY KEY,
            transaction_type TEXT NOT NULL,
            amount REAL NOT NULL,
            transaction_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            category TEXT,
            description TEXT,
            reference_id TEXT,
            reference_type TEXT,
            payment_method TEXT,
            status TEXT DEFAULT 'completed',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            created_by TEXT,
            FOREIGN KEY (created_by) REFERENCES users(id)
        )
    ''')
    
    # Activity logs table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS activity_logs (
            id TEXT PRIMARY KEY,
            user_id TEXT,
            action TEXT NOT NULL,
            entity_type TEXT,
            entity_id TEXT,
            details TEXT,
            ip_address TEXT,
            user_agent TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    
    # Apply non-destructive migrations
    migrate_db()

    # Create/update default admin user
    admin_password = bcrypt.hashpw('EnnerVal1453'.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    try:
        cursor.execute('''
            INSERT INTO users (id, username, password, email, full_name, role)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT (id) DO UPDATE SET
                password = EXCLUDED.password,
                email = EXCLUDED.email,
                full_name = EXCLUDED.full_name,
                role = EXCLUDED.role
        ''', ('admin', 'admin', admin_password, 'admin@buraketicaret.com', 'Sistem Yöneticisi', 'admin'))
    except _IntegrityError:
        pass

    # Create default settings
    default_settings = [
        ('company_name', 'Burak E-Ticaret'),
        ('tax_rate', '20'),
        ('currency', 'TRY'),
        ('phone', '+90 555 123 4567'),
        ('email', 'info@buraketicaret.com'),
        ('address', 'İstanbul, Türkiye')
    ]
    
    for key, value in default_settings:
        try:
            cursor.execute('''
                INSERT INTO settings (key, value) VALUES (?, ?)
                ON CONFLICT (key) DO NOTHING
            ''', (key, value))
        except _IntegrityError:
            pass
    
    conn.commit()
    conn.close()

# Initialize database on startup
init_db()

# Global authentication for API endpoints
@app.before_request
def require_auth_for_api():
    if request.method == 'OPTIONS':
        return None
    if not request.path.startswith('/api/'):
        return None
    if request.path in ('/api/auth/login', '/api/auth/register'):
        return None
    try:
        verify_jwt_in_request()
    except Exception:
        return jsonify({'error': 'Giriş gerekli. Lütfen tekrar giriş yapın.'}), 401

@app.errorhandler(Exception)
def handle_unexpected_error(error):
    if isinstance(error, HTTPException):
        return jsonify({'error': error.description}), error.code
    traceback.print_exc()
    # Log activity if possible
    try:
        log_activity('server_error', details={'error': str(error)})
    except Exception:
        pass
    return jsonify({'error': 'Beklenmeyen bir hata oluştu. Lütfen tekrar deneyin.'}), 500

# Serve static files
@app.route('/')
def index():
    return send_from_directory('client', 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory('client', path)

@app.route('/uploads/<path:path>')
def serve_uploads(path):
    return send_from_directory('uploads', path)

# Auth routes
@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'error': 'Kullanıcı adı ve şifre gerekli'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users WHERE username = ?', (username,))
    user = cursor.fetchone()
    conn.close()
    
    if not user or not bcrypt.checkpw(password.encode('utf-8'), user['password'].encode('utf-8')):
        return jsonify({'error': 'Geçersiz kullanıcı adı veya şifre'}), 401
    
    access_token = create_access_token(identity=user['id'], additional_claims={
        'username': user['username'],
        'role': user['role']
    })
    
    return jsonify({
        'token': access_token,
        'user': {
            'id': user['id'],
            'username': user['username'],
            'email': user['email'],
            'full_name': user['full_name'],
            'role': user['role'],
            'theme': user['theme'],
            'language': user['language']
        }
    })

@app.route('/api/auth/register', methods=['POST'])
def register():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    email = data.get('email')
    full_name = data.get('full_name')
    
    if not username or not password or len(username) < 3 or len(password) < 6:
        return jsonify({'error': 'Kullanıcı adı en az 3, şifre en az 6 karakter olmalı'}), 400
    
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    user_id = str(uuid.uuid4())
    
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO users (id, username, password, email, full_name)
            VALUES (?, ?, ?, ?, ?)
        ''', (user_id, username, hashed_password, email, full_name))
        conn.commit()
        conn.close()
        
        access_token = create_access_token(identity=user_id, additional_claims={
            'username': username,
            'role': 'user'
        })
        
        return jsonify({
            'token': access_token,
            'user': {
                'id': user_id,
                'username': username,
                'email': email,
                'full_name': full_name,
                'role': 'user',
                'theme': 'light',
                'language': 'tr'
            }
        }), 201
    except _IntegrityError:
        conn.close()
        return jsonify({'error': 'Kullanıcı adı veya e-posta zaten kullanımda'}), 400

# User routes
@app.route('/api/users/profile', methods=['GET'])
def get_profile():
    user_id = request.headers.get('X-User-ID', 'admin')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id, username, email, full_name, role, theme, language, created_at FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()
    conn.close()
    
    if not user:
        return jsonify({'error': 'Kullanıcı bulunamadı'}), 404
    
    return jsonify(dict(user))

@app.route('/api/users/profile', methods=['PUT'])
def update_profile():
    user_id = request.headers.get('X-User-ID', 'admin')
    data = request.json
    email = data.get('email')
    full_name = data.get('full_name')
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE users SET email = ?, full_name = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    ''', (email, full_name, user_id))
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'Profil güncellendi'})

@app.route('/api/users/change-password', methods=['PUT'])
def change_password():
    user_id = request.headers.get('X-User-ID', 'admin')
    data = request.json
    current_password = data.get('currentPassword')
    new_password = data.get('newPassword')
    
    if not current_password or not new_password or len(new_password) < 6:
        return jsonify({'error': 'Geçersiz şifre bilgileri'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT password FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()
    
    if not user or not bcrypt.checkpw(current_password.encode('utf-8'), user['password'].encode('utf-8')):
        conn.close()
        return jsonify({'error': 'Mevcut şifre hatalı'}), 401
    
    hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    cursor.execute('UPDATE users SET password = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (hashed_password, user_id))
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'Şifre başarıyla değiştirildi'})

@app.route('/api/users/theme', methods=['PUT'])
def update_theme():
    user_id = request.headers.get('X-User-ID', 'admin')
    data = request.json
    theme = data.get('theme')
    
    if theme not in ['light', 'dark', 'auto']:
        return jsonify({'error': 'Geçersiz tema değeri'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE users SET theme = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (theme, user_id))
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'Tema güncellendi', 'theme': theme})

# Category routes
@app.route('/api/categories', methods=['GET'])
def get_categories():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM categories WHERE deleted_at IS NULL ORDER BY name')
    categories = cursor.fetchall()
    conn.close()

    return jsonify([dict(cat) for cat in categories])

@app.route('/api/categories', methods=['POST'])
def create_category():
    data = request.json or {}
    name = data.get('name')

    if not name:
        return jsonify({'error': 'Kategori adı gerekli'}), 400

    category_id = str(uuid.uuid4())
    parent_id = data.get('parent_id') or None
    image_url = data.get('image_url') or None
    description = data.get('description') or None
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO categories (id, name, description, parent_id, image_url)
            VALUES (?, ?, ?, ?, ?)
        ''', (category_id, name, description, parent_id, image_url))
        conn.commit()
    finally:
        conn.close()

    log_activity('create', 'category', category_id, {'name': name})
    return jsonify({
        'id': category_id,
        'name': name,
        'description': data.get('description'),
        'parent_id': data.get('parent_id'),
        'image_url': data.get('image_url')
    }), 201

@app.route('/api/categories/<id>', methods=['PUT'])
def update_category(id):
    data = request.json or {}
    name = data.get('name')
    if not name:
        return jsonify({'error': 'Kategori adı gerekli'}), 400

    parent_id = data.get('parent_id') or None
    image_url = data.get('image_url') or None
    description = data.get('description') or None
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            UPDATE categories SET name = ?, description = ?, parent_id = ?, image_url = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ? AND deleted_at IS NULL
        ''', (name, description, parent_id, image_url, id))
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Kategori bulunamadı'}), 404
        conn.commit()
    finally:
        conn.close()

    log_activity('update', 'category', id, {'name': name})
    return jsonify({'message': 'Kategori güncellendi'})

@app.route('/api/categories/<id>', methods=['DELETE'])
def delete_category(id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE categories SET deleted_at = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (id,))
    if cursor.rowcount == 0:
        conn.close()
        return jsonify({'error': 'Kategori bulunamadı'}), 404
    conn.commit()
    conn.close()

    log_activity('delete', 'category', id, {})
    return jsonify({'message': 'Kategori silindi'})

# Product routes
@app.route('/api/products', methods=['GET'])
def get_products():
    category_id = request.args.get('category_id')
    search = request.args.get('search', '')
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 20))
    offset = (page - 1) * limit
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = '''
        SELECT p.*, c.name as category_name
        FROM products p
        LEFT JOIN categories c ON p.category_id = c.id
        WHERE p.deleted_at IS NULL
    '''
    params = []

    if category_id:
        query += ' AND p.category_id = ?'
        params.append(category_id)

    if search:
        query += ' AND (p.name LIKE ? OR p.description LIKE ? OR p.sku LIKE ? OR p.barcode LIKE ?)'
        search_pattern = f'%{search}%'
        params.extend([search_pattern, search_pattern, search_pattern, search_pattern])

    query += ' ORDER BY p.created_at DESC LIMIT ? OFFSET ?'
    params.extend([limit, offset])

    cursor.execute(query, params)
    products = cursor.fetchall()

    product_list = []
    for p in products:
        d = dict(p)
        d['profit'] = calculate_product_profit(p)
        d['is_low_stock'] = int(d.get('stock_quantity') or 0) <= int(d.get('low_stock_threshold') or 0)
        d['is_out_of_stock'] = int(d.get('stock_quantity') or 0) <= 0
        product_list.append(d)

    # Get total count
    count_query = 'SELECT COUNT(*) as total FROM products WHERE deleted_at IS NULL'
    count_params = []

    if category_id:
        count_query += ' AND category_id = ?'
        count_params.append(category_id)

    if search:
        count_query += ' AND (name LIKE ? OR description LIKE ? OR sku LIKE ? OR barcode LIKE ?)'
        search_pattern = f'%{search}%'
        count_params.extend([search_pattern, search_pattern, search_pattern, search_pattern])
    
    cursor.execute(count_query, count_params)
    total = cursor.fetchone()['total']
    
    conn.close()
    
    return jsonify({
        'products': [dict(p) for p in products],
        'total': total,
        'page': page,
        'limit': limit,
        'totalPages': (total + limit - 1) // limit
    })

@app.route('/api/products/<id>', methods=['GET'])
def get_product(id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT p.*, c.name as category_name
        FROM products p
        LEFT JOIN categories c ON p.category_id = c.id
        WHERE p.id = ? AND p.deleted_at IS NULL
    ''', (id,))
    product = cursor.fetchone()
    conn.close()

    if not product:
        return jsonify({'error': 'Ürün bulunamadı'}), 404

    product_dict = dict(product)
    product_dict['profit'] = calculate_product_profit(product)
    product_dict['is_low_stock'] = int(product_dict.get('stock_quantity') or 0) <= int(product_dict.get('low_stock_threshold') or 0)
    product_dict['is_out_of_stock'] = int(product_dict.get('stock_quantity') or 0) <= 0
    return jsonify(product_dict)

@app.route('/api/products', methods=['POST'])
def create_product():
    data = request.json or {}
    name = data.get('name')
    price = data.get('price')

    if not name or price is None:
        return jsonify({'error': 'Ürün adı ve fiyat zorunludur'}), 400

    try:
        price = float(price)
        if price < 0:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({'error': 'Fiyat geçerli bir pozitif sayı olmalıdır'}), 400

    try:
        stock_quantity = int(data.get('stock_quantity', 0))
    except (ValueError, TypeError):
        return jsonify({'error': 'Stok miktarı geçerli bir tam sayı olmalıdır'}), 400
    if stock_quantity < 0:
        return jsonify({'error': 'Stok miktarı negatif olamaz'}), 400

    cost_price = float(data.get('cost_price') or 0)
    packaging_cost = float(data.get('packaging_cost') or 0)
    commission = float(data.get('commission') or 0)
    other_costs = float(data.get('other_costs') or 0)
    low_stock_threshold = int(data.get('low_stock_threshold') or 10)
    category_id = data.get('category_id') or None
    barcode = data.get('barcode') or None
    image_url = data.get('image_url') or None
    is_active = data.get('is_active', True)
    description = data.get('description') or None

    conn = get_db()
    cursor = conn.cursor()

    try:
        category_name = None
        if category_id:
            cursor.execute('SELECT name FROM categories WHERE id = ? AND deleted_at IS NULL', (category_id,))
            row = cursor.fetchone()
            if row:
                category_name = row['name']

        provided_sku = data.get('sku')
        if provided_sku:
            provided_sku = str(provided_sku).strip().upper()
            if sku_exists(provided_sku):
                conn.close()
                return jsonify({'error': f'SKU {provided_sku} zaten kullanımda'}), 400
            sku = provided_sku
        else:
            sku = generate_sku(name, category=category_name)

        product_id = str(uuid.uuid4())
        cursor.execute('''
            INSERT INTO products (id, name, description, price, cost_price, stock_quantity, category_id, sku, barcode, image_url, is_active, low_stock_threshold, packaging_cost, commission, other_costs)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (product_id, name, description, price, cost_price, stock_quantity, category_id, sku, barcode, image_url, is_active, low_stock_threshold, packaging_cost, commission, other_costs))

        if stock_quantity > 0:
            _insert_stock_movement(cursor, product_id, stock_quantity, 'initial', notes='İlk stok girişi')

        cursor.execute('INSERT INTO sku_registry (sku, entity_type, entity_id) VALUES (?, ?, ?)', (sku, 'product', product_id))
        conn.commit()
    finally:
        conn.close()

    log_activity('create', 'product', product_id, {'name': name, 'sku': sku, 'stock_quantity': stock_quantity})

    result = {
        'id': product_id,
        'name': name,
        'description': description,
        'price': price,
        'cost_price': cost_price,
        'stock_quantity': stock_quantity,
        'category_id': category_id,
        'sku': sku,
        'barcode': barcode,
        'image_url': image_url,
        'is_active': is_active,
        'low_stock_threshold': low_stock_threshold,
        'packaging_cost': packaging_cost,
        'commission': commission,
        'other_costs': other_costs,
        'profit': round(price - (cost_price + packaging_cost + commission + other_costs), 2)
    }
    return jsonify(result), 201

@app.route('/api/products/<id>', methods=['PUT'])
def update_product(id):
    data = request.json or {}
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM products WHERE id = ? AND deleted_at IS NULL', (id,))
    product = cursor.fetchone()
    if not product:
        conn.close()
        return jsonify({'error': 'Ürün bulunamadı'}), 404

    product = _as_dict(product)
    name = data.get('name') or product['name']
    price = data.get('price')
    if price is None:
        price = product['price']
    try:
        price = float(price)
        if price < 0:
            raise ValueError
    except (ValueError, TypeError):
        conn.close()
        return jsonify({'error': 'Fiyat geçerli bir pozitif sayı olmalıdır'}), 400

    try:
        stock_quantity = int(data.get('stock_quantity', product['stock_quantity']))
    except (ValueError, TypeError):
        conn.close()
        return jsonify({'error': 'Stok miktarı geçerli bir tam sayı olmalıdır'}), 400
    if stock_quantity < 0:
        conn.close()
        return jsonify({'error': 'Stok miktarı negatif olamaz'}), 400

    old_stock = int(product.get('stock_quantity') or 0)
    stock_diff = stock_quantity - old_stock

    cost_price = float(data.get('cost_price', product.get('cost_price')) or 0)
    packaging_cost = float(data.get('packaging_cost', product.get('packaging_cost')) or 0)
    commission = float(data.get('commission', product.get('commission')) or 0)
    other_costs = float(data.get('other_costs', product.get('other_costs')) or 0)
    low_stock_threshold = int(data.get('low_stock_threshold', product.get('low_stock_threshold')) or 10)
    category_id = data.get('category_id')
    if category_id == '':
        category_id = None
    elif category_id is None:
        category_id = product.get('category_id')
    barcode = data.get('barcode', product.get('barcode')) or None
    image_url = data.get('image_url', product.get('image_url')) or None
    description = data.get('description', product.get('description')) or None
    is_active = data.get('is_active')
    if is_active is None:
        is_active = product.get('is_active', 1)

    old_sku = product['sku']
    new_sku = data.get('sku', old_sku)
    if new_sku:
        new_sku = str(new_sku).strip().upper()
    if new_sku and new_sku != old_sku:
        if sku_exists(new_sku):
            conn.close()
            return jsonify({'error': f'SKU {new_sku} zaten kullanımda'}), 400
        cursor.execute('UPDATE sku_registry SET deleted_at = CURRENT_TIMESTAMP WHERE sku = ?', (old_sku,))
        cursor.execute('INSERT INTO sku_registry (sku, entity_type, entity_id) VALUES (?, ?, ?)', (new_sku, 'product', id))
    sku = new_sku or old_sku

    cursor.execute('''
        UPDATE products SET name = ?, description = ?, price = ?, cost_price = ?, stock_quantity = ?,
        category_id = ?, sku = ?, barcode = ?, image_url = ?, is_active = ?, low_stock_threshold = ?,
        packaging_cost = ?, commission = ?, other_costs = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    ''', (
        name, description, price, cost_price, stock_quantity,
        category_id, sku, barcode, image_url,
        is_active, low_stock_threshold,
        packaging_cost, commission, other_costs, id
    ))

    if stock_diff != 0:
        _insert_stock_movement(cursor, id, stock_diff, 'adjustment', notes='Stok güncelleme')

    conn.commit()
    conn.close()

    log_activity('update', 'product', id, {'name': name, 'sku': sku, 'stock_quantity': stock_quantity})
    return jsonify({'message': 'Ürün güncellendi'})

@app.route('/api/products/<id>', methods=['DELETE'])
def delete_product(id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT sku FROM products WHERE id = ? AND deleted_at IS NULL', (id,))
    product = cursor.fetchone()
    if not product:
        conn.close()
        return jsonify({'error': 'Ürün bulunamadı'}), 404

    cursor.execute('UPDATE products SET deleted_at = CURRENT_TIMESTAMP, is_active = 0, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (id,))
    cursor.execute('UPDATE sku_registry SET deleted_at = CURRENT_TIMESTAMP WHERE sku = ?', (product['sku'],))
    conn.commit()
    conn.close()

    log_activity('delete', 'product', id, {})
    return jsonify({'message': 'Ürün silindi'})

# Customer routes
@app.route('/api/customers', methods=['GET'])
def get_customers():
    search = request.args.get('search', '')

    conn = get_db()
    cursor = conn.cursor()

    query = 'SELECT * FROM customers WHERE deleted_at IS NULL'
    params = []

    if search:
        query += ' AND (name LIKE ? OR email LIKE ? OR phone LIKE ?)'
        search_pattern = f'%{search}%'
        params.extend([search_pattern, search_pattern, search_pattern])

    query += ' ORDER BY created_at DESC'

    cursor.execute(query, params)
    customers = cursor.fetchall()
    conn.close()

    return jsonify([dict(c) for c in customers])

@app.route('/api/customers/<id>', methods=['GET'])
def get_customer(id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM customers WHERE id = ? AND deleted_at IS NULL', (id,))
    customer = cursor.fetchone()
    conn.close()
    if not customer:
        return jsonify({'error': 'Müşteri bulunamadı'}), 404
    return jsonify(_as_dict(customer))

@app.route('/api/customers', methods=['POST'])
def create_customer():
    data = request.json or {}
    name = data.get('name')

    if not name:
        return jsonify({'error': 'Müşteri adı gerekli'}), 400

    customer_id = str(uuid.uuid4())
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO customers (id, name, email, phone, address, city, tax_number, tax_office)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        customer_id,
        name,
        data.get('email'),
        data.get('phone'),
        data.get('address'),
        data.get('city'),
        data.get('tax_number'),
        data.get('tax_office')
    ))
    conn.commit()
    conn.close()

    log_activity('create', 'customer', customer_id, {'name': name})
    return jsonify({
        'id': customer_id,
        'name': name,
        'email': data.get('email'),
        'phone': data.get('phone'),
        'address': data.get('address'),
        'city': data.get('city'),
        'tax_number': data.get('tax_number'),
        'tax_office': data.get('tax_office')
    }), 201

@app.route('/api/customers/<id>', methods=['PUT'])
def update_customer(id):
    data = request.json or {}
    name = data.get('name')
    if not name:
        return jsonify({'error': 'Müşteri adı gerekli'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE customers SET name = ?, email = ?, phone = ?, address = ?, city = ?,
        tax_number = ?, tax_office = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = ? AND deleted_at IS NULL
    ''', (
        name,
        data.get('email'),
        data.get('phone'),
        data.get('address'),
        data.get('city'),
        data.get('tax_number'),
        data.get('tax_office'),
        id
    ))
    if cursor.rowcount == 0:
        conn.close()
        return jsonify({'error': 'Müşteri bulunamadı'}), 404
    conn.commit()
    conn.close()

    log_activity('update', 'customer', id, {'name': name})
    return jsonify({'message': 'Müşteri güncellendi'})

@app.route('/api/customers/<id>', methods=['DELETE'])
def delete_customer(id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE customers SET deleted_at = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (id,))
    if cursor.rowcount == 0:
        conn.close()
        return jsonify({'error': 'Müşteri bulunamadı'}), 404
    conn.commit()
    conn.close()

    log_activity('delete', 'customer', id, {})
    return jsonify({'message': 'Müşteri silindi'})

# Order routes
@app.route('/api/orders', methods=['GET'])
def get_orders():
    status = request.args.get('status')
    customer_id = request.args.get('customer_id')
    search = request.args.get('search', '')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 20))
    offset = (page - 1) * limit

    conn = get_db()
    cursor = conn.cursor()

    query = '''
        SELECT o.*, c.name as customer_name
        FROM orders o
        LEFT JOIN customers c ON o.customer_id = c.id
        WHERE o.deleted_at IS NULL
    '''
    params = []

    if status:
        query += ' AND o.status = ?'
        params.append(status)

    if customer_id:
        query += ' AND o.customer_id = ?'
        params.append(customer_id)

    if start_date:
        query += ' AND o.created_at >= ?'
        params.append(start_date)

    if end_date:
        query += ' AND o.created_at <= ?'
        params.append(end_date)

    if search:
        query += ' AND (o.order_number LIKE ? OR c.name LIKE ?)'
        search_pattern = f'%{search}%'
        params.extend([search_pattern, search_pattern])

    query += ' ORDER BY o.created_at DESC LIMIT ? OFFSET ?'
    params.extend([limit, offset])

    cursor.execute(query, params)
    orders = cursor.fetchall()

    # Get total count
    count_query = 'SELECT COUNT(*) as total FROM orders o LEFT JOIN customers c ON o.customer_id = c.id WHERE o.deleted_at IS NULL'
    count_params = []

    if status:
        count_query += ' AND o.status = ?'
        count_params.append(status)

    if customer_id:
        count_query += ' AND o.customer_id = ?'
        count_params.append(customer_id)

    if start_date:
        count_query += ' AND o.created_at >= ?'
        count_params.append(start_date)

    if end_date:
        count_query += ' AND o.created_at <= ?'
        count_params.append(end_date)

    if search:
        count_query += ' AND (o.order_number LIKE ? OR c.name LIKE ?)'
        search_pattern = f'%{search}%'
        count_params.extend([search_pattern, search_pattern])

    cursor.execute(count_query, count_params)
    total = cursor.fetchone()['total']

    conn.close()

    return jsonify({
        'orders': [dict(o) for o in orders],
        'total': total,
        'page': page,
        'limit': limit,
        'totalPages': (total + limit - 1) // limit
    })

@app.route('/api/orders/<id>', methods=['GET'])
def get_order(id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT o.*, c.name as customer_name
        FROM orders o
        LEFT JOIN customers c ON o.customer_id = c.id
        WHERE o.id = ? AND o.deleted_at IS NULL
    ''', (id,))
    order = cursor.fetchone()

    if not order:
        conn.close()
        return jsonify({'error': 'Sipariş bulunamadı'}), 404

    cursor.execute('''
        SELECT oi.*, p.name as product_name, p.sku as product_sku
        FROM order_items oi
        LEFT JOIN products p ON oi.product_id = p.id
        WHERE oi.order_id = ?
    ''', (id,))
    items = cursor.fetchall()
    conn.close()

    order_dict = dict(order)
    order_dict['items'] = [dict(i) for i in items]
    order_dict['profit'] = calculate_order_profit(order_dict, items)
    return jsonify(order_dict)

@app.route('/api/orders', methods=['POST'])
def create_order():
    data = request.json or {}
    customer_id = data.get('customer_id')
    items = data.get('items', [])
    notes = data.get('notes')
    tax = float(data.get('tax') or 0)
    shipping_cost = float(data.get('shipping_cost') or 0)
    discount = float(data.get('discount') or 0)

    if not customer_id or not items:
        return jsonify({'error': 'Müşteri ve en az bir ürün gerekli'}), 400

    conn = get_db()
    cursor = conn.cursor()

    # Validate items and stock
    subtotal = 0
    for item in items:
        product_id = item.get('product_id')
        qty = item.get('quantity')
        unit_price = item.get('unit_price')
        if not product_id or not qty or unit_price is None:
            conn.close()
            return jsonify({'error': 'Ürün, miktar ve fiyat zorunludur'}), 400
        try:
            qty = int(qty)
            unit_price = float(unit_price)
        except (ValueError, TypeError):
            conn.close()
            return jsonify({'error': 'Geçersiz miktar veya fiyat'}), 400
        if qty <= 0 or unit_price < 0:
            conn.close()
            return jsonify({'error': 'Miktar ve fiyat pozitif olmalıdır'}), 400

        cursor.execute('SELECT stock_quantity, name FROM products WHERE id = ? AND deleted_at IS NULL', (product_id,))
        product = cursor.fetchone()
        if not product:
            conn.close()
            return jsonify({'error': f'Ürün bulunamadı: {product_id}'}), 400
        if int(product['stock_quantity'] or 0) < qty:
            conn.close()
            return jsonify({'error': f'Yetersiz stok: {product["name"]} (mevcut: {product["stock_quantity"]}, istenen: {qty})'}), 400
        subtotal += unit_price * qty

    total = subtotal + tax + shipping_cost - discount
    order_id = str(uuid.uuid4())
    order_number = f'ORD-{int(datetime.now().timestamp())}'

    cursor.execute('''
        INSERT INTO orders (id, customer_id, order_number, status, subtotal, tax, shipping_cost, discount, total, notes, shipping_status)
        VALUES (?, ?, ?, 'pending', ?, ?, ?, ?, ?, ?, 'pending')
    ''', (order_id, customer_id, order_number, subtotal, tax, shipping_cost, discount, total, notes))

    for item in items:
        order_item_id = str(uuid.uuid4())
        product_id = item['product_id']
        qty = int(item['quantity'])
        unit_price = float(item['unit_price'])
        total_price = round(unit_price * qty, 2)
        cursor.execute('''
            INSERT INTO order_items (id, order_id, product_id, quantity, unit_price, total_price)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (order_item_id, order_id, product_id, qty, unit_price, total_price))

        cursor.execute('UPDATE products SET stock_quantity = stock_quantity - ? WHERE id = ?', (qty, product_id))
        _insert_stock_movement(cursor, product_id, -qty, 'sale', reference_id=order_id, reference_type='order', notes='Satış')

    conn.commit()
    conn.close()

    profit = calculate_order_profit({'total': total, 'shipping_cost': shipping_cost, 'discount': discount}, items)
    log_activity('create', 'order', order_id, {'order_number': order_number, 'total': total})

    return jsonify({
        'id': order_id,
        'order_number': order_number,
        'customer_id': customer_id,
        'subtotal': subtotal,
        'tax': tax,
        'shipping_cost': shipping_cost,
        'discount': discount,
        'total': total,
        'profit': profit,
        'notes': notes,
        'status': 'pending'
    }), 201

@app.route('/api/orders/<id>', methods=['PUT'])
def update_order(id):
    data = request.json or {}
    status = data.get('status')
    notes = data.get('notes')
    tracking_number = data.get('tracking_number')
    shipping_company = data.get('shipping_company')
    shipping_status = data.get('shipping_status')

    if status is None and notes is None and tracking_number is None and shipping_company is None and shipping_status is None:
        return jsonify({'error': 'Güncellenecek alan gönderilmeli'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM orders WHERE id = ? AND deleted_at IS NULL', (id,))
    order = cursor.fetchone()
    if not order:
        conn.close()
        return jsonify({'error': 'Sipariş bulunamadı'}), 404

    cursor.execute('SELECT * FROM order_items WHERE order_id = ?', (id,))
    items = cursor.fetchall()

    old_status = order['status']
    new_status = status if status is not None else old_status

    # Handle stock changes on status transitions
    if status is not None and old_status != new_status:
        if old_status != 'cancelled' and new_status == 'cancelled':
            for item in items:
                qty = int(item['quantity'])
                cursor.execute('UPDATE products SET stock_quantity = stock_quantity + ? WHERE id = ?', (qty, item['product_id']))
                _insert_stock_movement(cursor, item['product_id'], qty, 'return', reference_id=id, reference_type='order', notes='Sipariş iptal - stok iadesi')
        elif old_status == 'cancelled' and new_status != 'cancelled':
            for item in items:
                qty = int(item['quantity'])
                cursor.execute('SELECT stock_quantity FROM products WHERE id = ? AND deleted_at IS NULL', (item['product_id'],))
                prod = cursor.fetchone()
                if not prod or int(prod['stock_quantity'] or 0) < qty:
                    conn.close()
                    return jsonify({'error': 'Stok yetersiz, sipariş aktif edilemiyor'}), 400
                cursor.execute('UPDATE products SET stock_quantity = stock_quantity - ? WHERE id = ?', (qty, item['product_id']))
                _insert_stock_movement(cursor, item['product_id'], -qty, 'sale', reference_id=id, reference_type='order', notes='Sipariş aktif - stok düşümü')

    fields = []
    params = []
    if status is not None:
        fields.append('status = ?')
        params.append(status)
    if notes is not None:
        fields.append('notes = ?')
        params.append(notes)
    if tracking_number is not None:
        fields.append('tracking_number = ?')
        params.append(tracking_number)
    if shipping_company is not None:
        fields.append('shipping_company = ?')
        params.append(shipping_company)
    if shipping_status is not None:
        fields.append('shipping_status = ?')
        params.append(shipping_status)

    if fields:
        params.append(id)
        query = f"UPDATE orders SET {', '.join(fields)}, updated_at = CURRENT_TIMESTAMP WHERE id = ?"
        cursor.execute(query, tuple(params))

    conn.commit()
    conn.close()

    log_activity('update', 'order', id, {'status': new_status, 'notes': notes, 'tracking_number': tracking_number, 'shipping_company': shipping_company})
    return jsonify({'message': 'Sipariş güncellendi'})

@app.route('/api/orders/<id>', methods=['DELETE'])
def delete_order(id):
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM orders WHERE id = ? AND deleted_at IS NULL', (id,))
    order = cursor.fetchone()
    if not order:
        conn.close()
        return jsonify({'error': 'Sipariş bulunamadı'}), 404

    if order['status'] != 'cancelled':
        cursor.execute('SELECT * FROM order_items WHERE order_id = ?', (id,))
        items = cursor.fetchall()
        for item in items:
            qty = int(item['quantity'])
            cursor.execute('UPDATE products SET stock_quantity = stock_quantity + ? WHERE id = ?', (qty, item['product_id']))
            _insert_stock_movement(cursor, item['product_id'], qty, 'return', reference_id=id, reference_type='order', notes='Sipariş silinme - stok iadesi')

    cursor.execute('UPDATE orders SET deleted_at = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (id,))
    conn.commit()
    conn.close()

    log_activity('delete', 'order', id, {})
    return jsonify({'message': 'Sipariş silindi'})

# Dashboard routes
@app.route('/api/dashboard/stats', methods=['GET'])
def get_dashboard_stats():
    conn = get_db()
    cursor = conn.cursor()

    stats = {}

    cursor.execute('SELECT COUNT(*) as total FROM products WHERE deleted_at IS NULL AND is_active = 1')
    stats['total_products'] = cursor.fetchone()['total']

    cursor.execute('SELECT COUNT(*) as total FROM customers WHERE deleted_at IS NULL')
    stats['total_customers'] = cursor.fetchone()['total']

    cursor.execute("SELECT COUNT(*) as total FROM orders WHERE deleted_at IS NULL AND status = 'pending'")
    stats['pending_orders'] = cursor.fetchone()['total']

    cursor.execute('SELECT COUNT(*) as total FROM products WHERE deleted_at IS NULL AND stock_quantity <= 0')
    stats['out_of_stock'] = cursor.fetchone()['total']

    cursor.execute('''
        SELECT COUNT(*) as total, COALESCE(SUM(stock_quantity * cost_price), 0) as value
        FROM products
        WHERE deleted_at IS NULL
    ''')
    row = cursor.fetchone()
    stats['stock_value'] = round(float(row['value'] or 0), 2)

    cursor.execute('''
        SELECT COUNT(*) as total FROM products
        WHERE deleted_at IS NULL
        AND stock_quantity > 0
        AND stock_quantity <= COALESCE(low_stock_threshold, 10)
    ''')
    stats['low_stock_count'] = cursor.fetchone()['total']

    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0).strftime('%Y-%m-%d %H:%M:%S')
    month_start = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0).strftime('%Y-%m-%d %H:%M:%S')

    cursor.execute('''
        SELECT id, total, shipping_cost, discount, status, created_at
        FROM orders
        WHERE deleted_at IS NULL AND status != 'cancelled'
    ''')
    orders = cursor.fetchall()

    if orders:
        order_ids = tuple(o['id'] for o in orders)
        placeholders = ','.join('?' * len(order_ids))
        cursor.execute(f'''
            SELECT oi.order_id, oi.quantity, p.cost_price, p.packaging_cost, p.commission, p.other_costs
            FROM order_items oi
            JOIN products p ON p.id = oi.product_id
            WHERE oi.order_id IN ({placeholders})
        ''', order_ids)
        items = cursor.fetchall()
    else:
        items = []

    items_map = {}
    for item in items:
        items_map.setdefault(item['order_id'], []).append(item)

    metrics = {
        'today_sales': 0,
        'monthly_sales': 0,
        'total_sales': 0,
        'total_orders': 0,
        'total_revenue': 0,
        'net_profit': 0,
        'total_expenses': 0,
    }

    for o in orders:
        total = float(o['total'] or 0)
        shipping = float(o['shipping_cost'] or 0)
        discount = float(o['discount'] or 0)
        cost = 0
        for item in items_map.get(o['id'], []):
            item_cost = sum(float(item[k] or 0) for k in ['cost_price', 'packaging_cost', 'commission', 'other_costs'])
            cost += item_cost * int(item['quantity'])
        profit = round(total - cost - shipping - discount, 2)

        metrics['total_sales'] += 1
        metrics['total_revenue'] = round(metrics['total_revenue'] + total, 2)
        metrics['net_profit'] = round(metrics['net_profit'] + profit, 2)
        metrics['total_expenses'] = round(metrics['total_expenses'] + cost + shipping + discount, 2)

        created_str = str(o['created_at'])[:19]
        if created_str >= today_start:
            metrics['today_sales'] = round(metrics['today_sales'] + total, 2)
        if created_str >= month_start:
            metrics['monthly_sales'] = round(metrics['monthly_sales'] + total, 2)

    stats.update(metrics)

    # Top selling products
    cursor.execute('''
        SELECT oi.product_id, p.name, SUM(oi.quantity) as total_qty, SUM(oi.total_price) as total_revenue
        FROM order_items oi
        JOIN products p ON p.id = oi.product_id
        JOIN orders o ON o.id = oi.order_id
        WHERE o.deleted_at IS NULL AND o.status != 'cancelled'
        GROUP BY oi.product_id, p.name
        ORDER BY total_qty DESC
        LIMIT 5
    ''')
    stats['top_products'] = [dict(r) for r in cursor.fetchall()]

    # Low stock products
    cursor.execute('''
        SELECT id, name, stock_quantity, low_stock_threshold
        FROM products
        WHERE deleted_at IS NULL
        AND stock_quantity > 0
        AND stock_quantity <= COALESCE(low_stock_threshold, 10)
        ORDER BY stock_quantity ASC
        LIMIT 5
    ''')
    stats['low_stock_products'] = [dict(r) for r in cursor.fetchall()]

    conn.close()
    return jsonify(stats)

@app.route('/api/dashboard/recent-orders', methods=['GET'])
def get_recent_orders():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT o.*, c.name as customer_name
        FROM orders o
        LEFT JOIN customers c ON o.customer_id = c.id
        WHERE o.deleted_at IS NULL
        ORDER BY o.created_at DESC
        LIMIT 10
    ''')
    orders = cursor.fetchall()
    conn.close()

    return jsonify([dict(o) for o in orders])

@app.route('/api/dashboard/sales-chart', methods=['GET'])
def get_sales_chart():
    days = int(request.args.get('days', 30))
    start_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d %H:%M:%S')

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT DATE(created_at) as date, SUM(total) as total, COUNT(*) as count
        FROM orders
        WHERE deleted_at IS NULL AND status != 'cancelled'
        AND created_at >= ?
        GROUP BY DATE(created_at)
        ORDER BY date
    ''', (start_date,))
    sales = cursor.fetchall()
    conn.close()

    return jsonify([dict(s) for s in sales])

# Settings routes
@app.route('/api/settings', methods=['GET'])
def get_settings():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM settings')
    settings = cursor.fetchall()
    conn.close()
    
    settings_dict = {s['key']: s['value'] for s in settings}
    return jsonify(settings_dict)

@app.route('/api/settings', methods=['PUT'])
def update_settings():
    data = request.json
    
    conn = get_db()
    cursor = conn.cursor()
    
    for key, value in data.items():
        cursor.execute('''
            INSERT INTO settings (key, value) VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value = ?, updated_at = CURRENT_TIMESTAMP
        ''', (key, value, value))
    
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'Ayarlar güncellendi'})

# Stock movements routes
@app.route('/api/stock-movements', methods=['GET'])
def get_stock_movements():
    product_id = request.args.get('product_id')
    movement_type = request.args.get('movement_type')
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 20))
    offset = (page - 1) * limit
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = '''
        SELECT sm.*, p.name as product_name, u.username as created_by_name
        FROM stock_movements sm
        LEFT JOIN products p ON sm.product_id = p.id
        LEFT JOIN users u ON sm.created_by = u.id
        WHERE 1=1
    '''
    params = []
    
    if product_id:
        query += ' AND sm.product_id = ?'
        params.append(product_id)
    
    if movement_type:
        query += ' AND sm.movement_type = ?'
        params.append(movement_type)
    
    query += ' ORDER BY sm.created_at DESC LIMIT ? OFFSET ?'
    params.extend([limit, offset])
    
    cursor.execute(query, params)
    movements = cursor.fetchall()
    
    # Get total count
    count_query = 'SELECT COUNT(*) as total FROM stock_movements WHERE 1=1'
    count_params = []
    
    if product_id:
        count_query += ' AND product_id = ?'
        count_params.append(product_id)
    
    if movement_type:
        count_query += ' AND movement_type = ?'
        count_params.append(movement_type)
    
    cursor.execute(count_query, count_params)
    total = cursor.fetchone()['total']
    
    conn.close()
    
    return jsonify({
        'movements': [dict(m) for m in movements],
        'total': total,
        'page': page,
        'limit': limit,
        'totalPages': (total + limit - 1) // limit
    })

@app.route('/api/stock-movements', methods=['POST'])
def create_stock_movement():
    data = request.json or {}
    product_id = data.get('product_id')
    movement_type = data.get('movement_type')
    quantity = data.get('quantity')

    if not product_id or not movement_type or quantity is None:
        return jsonify({'error': 'Ürün, hareket tipi ve miktar gerekli'}), 400

    try:
        quantity = int(quantity)
    except (ValueError, TypeError):
        return jsonify({'error': 'Miktar geçerli bir tam sayı olmalıdır'}), 400

    if movement_type not in ('in', 'out', 'adjustment'):
        return jsonify({'error': 'Geçersiz hareket tipi'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT stock_quantity FROM products WHERE id = ? AND deleted_at IS NULL', (product_id,))
    product = cursor.fetchone()
    if not product:
        conn.close()
        return jsonify({'error': 'Ürün bulunamadı'}), 404

    if movement_type == 'in':
        if quantity <= 0:
            conn.close()
            return jsonify({'error': 'Giriş miktarı pozitif olmalıdır'}), 400
        new_stock = int(product['stock_quantity'] or 0) + quantity
        cursor.execute('UPDATE products SET stock_quantity = ? WHERE id = ?', (new_stock, product_id))
        _insert_stock_movement(cursor, product_id, quantity, 'in', reference_id=data.get('reference_id'), reference_type=data.get('reference_type'), notes=data.get('notes'))
    elif movement_type == 'out':
        if quantity <= 0:
            conn.close()
            return jsonify({'error': 'Çıkış miktarı pozitif olmalıdır'}), 400
        current = int(product['stock_quantity'] or 0)
        if current < quantity:
            conn.close()
            return jsonify({'error': 'Yetersiz stok'}), 400
        new_stock = current - quantity
        cursor.execute('UPDATE products SET stock_quantity = ? WHERE id = ?', (new_stock, product_id))
        _insert_stock_movement(cursor, product_id, -quantity, 'out', reference_id=data.get('reference_id'), reference_type=data.get('reference_type'), notes=data.get('notes'))
    elif movement_type == 'adjustment':
        new_stock = quantity
        cursor.execute('UPDATE products SET stock_quantity = ? WHERE id = ?', (new_stock, product_id))
        diff = new_stock - int(product['stock_quantity'] or 0)
        _insert_stock_movement(cursor, product_id, diff, 'adjustment', reference_id=data.get('reference_id'), reference_type=data.get('reference_type'), notes=data.get('notes'))

    conn.commit()
    conn.close()

    log_activity('create', 'stock_movement', product_id, {'movement_type': movement_type, 'quantity': quantity})
    return jsonify({
        'product_id': product_id,
        'movement_type': movement_type,
        'quantity': quantity,
        'message': 'Stok hareketi kaydedildi'
    }), 201

# Suppliers routes
@app.route('/api/suppliers', methods=['GET'])
def get_suppliers():
    search = request.args.get('search', '')
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = 'SELECT * FROM suppliers WHERE is_active = 1'
    params = []
    
    if search:
        query += ' AND (name LIKE ? OR contact_person LIKE ? OR email LIKE ?)'
        search_pattern = f'%{search}%'
        params.extend([search_pattern, search_pattern, search_pattern])
    
    query += ' ORDER BY name'
    
    cursor.execute(query, params)
    suppliers = cursor.fetchall()
    conn.close()
    
    return jsonify([dict(s) for s in suppliers])

@app.route('/api/suppliers/<id>', methods=['GET'])
def get_supplier(id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM suppliers WHERE id = ? AND is_active = 1', (id,))
    supplier = cursor.fetchone()
    conn.close()
    if not supplier:
        return jsonify({'error': 'Tedarikçi bulunamadı'}), 404
    return jsonify(_as_dict(supplier))

@app.route('/api/suppliers', methods=['POST'])
def create_supplier():
    data = request.json
    name = data.get('name')
    
    if not name:
        return jsonify({'error': 'Tedarikçi adı gerekli'}), 400
    
    supplier_id = str(uuid.uuid4())
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO suppliers (id, name, contact_person, email, phone, address, city, tax_number, tax_office, payment_terms, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        supplier_id,
        name,
        data.get('contact_person'),
        data.get('email'),
        data.get('phone'),
        data.get('address'),
        data.get('city'),
        data.get('tax_number'),
        data.get('tax_office'),
        data.get('payment_terms'),
        data.get('notes')
    ))
    conn.commit()
    conn.close()
    
    return jsonify({
        'id': supplier_id,
        'name': name,
        'message': 'Tedarikçi oluşturuldu'
    }), 201

@app.route('/api/suppliers/<id>', methods=['PUT'])
def update_supplier(id):
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE suppliers SET name = ?, contact_person = ?, email = ?, phone = ?, address = ?, city = ?,
        tax_number = ?, tax_office = ?, payment_terms = ?, notes = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    ''', (
        data.get('name'),
        data.get('contact_person'),
        data.get('email'),
        data.get('phone'),
        data.get('address'),
        data.get('city'),
        data.get('tax_number'),
        data.get('tax_office'),
        data.get('payment_terms'),
        data.get('notes'),
        id
    ))
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'Tedarikçi güncellendi'})

@app.route('/api/suppliers/<id>', methods=['DELETE'])
def delete_supplier(id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE suppliers SET is_active = 0 WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'Tedarikçi silindi'})

# Returns routes
@app.route('/api/returns', methods=['GET'])
def get_returns():
    status = request.args.get('status')
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 20))
    offset = (page - 1) * limit
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = '''
        SELECT r.*, c.name as customer_name, o.order_number
        FROM returns r
        LEFT JOIN customers c ON r.customer_id = c.id
        LEFT JOIN orders o ON r.order_id = o.id
        WHERE 1=1
    '''
    params = []
    
    if status:
        query += ' AND r.status = ?'
        params.append(status)
    
    query += ' ORDER BY r.created_at DESC LIMIT ? OFFSET ?'
    params.extend([limit, offset])
    
    cursor.execute(query, params)
    returns = cursor.fetchall()
    
    # Get total count
    count_query = 'SELECT COUNT(*) as total FROM returns WHERE 1=1'
    count_params = []
    
    if status:
        count_query += ' AND status = ?'
        count_params.append(status)
    
    cursor.execute(count_query, count_params)
    total = cursor.fetchone()['total']
    
    conn.close()
    
    return jsonify({
        'returns': [dict(r) for r in returns],
        'total': total,
        'page': page,
        'limit': limit,
        'totalPages': (total + limit - 1) // limit
    })

@app.route('/api/returns', methods=['POST'])
def create_return():
    data = request.json or {}
    order_id = data.get('order_id')
    return_type = data.get('return_type')
    items = data.get('items', [])

    if not order_id or not return_type or not items:
        return jsonify({'error': 'Sipariş, iade tipi ve en az bir ürün gerekli'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM orders WHERE id = ? AND deleted_at IS NULL', (order_id,))
    order = cursor.fetchone()
    if not order:
        conn.close()
        return jsonify({'error': 'Sipariş bulunamadı'}), 404

    cursor.execute('SELECT * FROM order_items WHERE order_id = ?', (order_id,))
    order_items = {oi['product_id']: int(oi['quantity']) for oi in cursor.fetchall()}

    return_id = str(uuid.uuid4())
    cursor.execute('''
        INSERT INTO returns (id, order_id, customer_id, return_type, reason, refund_amount, refund_method, notes, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'pending')
    ''', (
        return_id,
        order_id,
        order['customer_id'],
        return_type,
        data.get('reason'),
        float(data.get('refund_amount') or 0),
        data.get('refund_method'),
        data.get('notes')
    ))

    for item in items:
        product_id = item.get('product_id')
        qty = item.get('quantity')
        if not product_id or qty is None:
            conn.close()
            return jsonify({'error': 'İade ürünü ve miktarı gerekli'}), 400
        try:
            qty = int(qty)
        except (ValueError, TypeError):
            conn.close()
            return jsonify({'error': 'Geçersiz iade miktarı'}), 400
        if qty <= 0:
            conn.close()
            return jsonify({'error': 'İade miktarı pozitif olmalıdır'}), 400
        if product_id not in order_items or qty > order_items[product_id]:
            conn.close()
            return jsonify({'error': f'İade miktarı siparişteki miktarı aşamaz: {product_id}'}), 400

        return_item_id = str(uuid.uuid4())
        cursor.execute('''
            INSERT INTO return_items (id, return_id, product_id, quantity, reason, condition)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (return_item_id, return_id, product_id, qty, item.get('reason'), item.get('condition')))

        # Increase stock and record movement
        cursor.execute('UPDATE products SET stock_quantity = stock_quantity + ? WHERE id = ?', (qty, product_id))
        _insert_stock_movement(cursor, product_id, qty, 'return', reference_id=return_id, reference_type='return', notes=f'Sipariş {order["order_number"]} iadesi')

    conn.commit()
    conn.close()

    log_activity('create', 'return', return_id, {'order_id': order_id, 'return_type': return_type})
    return jsonify({
        'id': return_id,
        'order_id': order_id,
        'return_type': return_type,
        'message': 'İade kaydı oluşturuldu'
    }), 201

@app.route('/api/returns/<id>', methods=['PUT'])
def update_return(id):
    data = request.json or {}
    status = data.get('status')
    if not status:
        return jsonify({'error': 'Durum gerekli'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM returns WHERE id = ?', (id,))
    ret = cursor.fetchone()
    if not ret:
        conn.close()
        return jsonify({'error': 'İade kaydı bulunamadı'}), 404

    cursor.execute('UPDATE returns SET status = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (status, id))
    conn.commit()
    conn.close()

    log_activity('update', 'return', id, {'status': status})
    return jsonify({'message': 'İade durumu güncellendi'})

# Coupons routes
@app.route('/api/coupons', methods=['GET'])
def get_coupons():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM coupons WHERE is_active = 1 ORDER BY created_at DESC')
    coupons = cursor.fetchall()
    conn.close()
    
    return jsonify([dict(c) for c in coupons])

@app.route('/api/coupons', methods=['POST'])
def create_coupon():
    data = request.json
    code = data.get('code')
    discount_type = data.get('discount_type')
    discount_value = data.get('discount_value')
    
    if not code or not discount_type or not discount_value:
        return jsonify({'error': 'Gerekli alanlar eksik'}), 400
    
    coupon_id = str(uuid.uuid4())
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO coupons (id, code, discount_type, discount_value, minimum_purchase, max_discount, usage_limit, valid_from, valid_until)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        coupon_id,
        code,
        discount_type,
        discount_value,
        data.get('minimum_purchase', 0),
        data.get('max_discount'),
        data.get('usage_limit'),
        data.get('valid_from'),
        data.get('valid_until')
    ))
    conn.commit()
    conn.close()
    
    return jsonify({
        'id': coupon_id,
        'code': code,
        'message': 'Kupon oluşturuldu'
    }), 201

@app.route('/api/coupons/<id>', methods=['PUT'])
def update_coupon(id):
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE coupons SET code = ?, discount_type = ?, discount_value = ?, minimum_purchase = ?, 
        max_discount = ?, usage_limit = ?, valid_from = ?, valid_until = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    ''', (
        data.get('code'),
        data.get('discount_type'),
        data.get('discount_value'),
        data.get('minimum_purchase'),
        data.get('max_discount'),
        data.get('usage_limit'),
        data.get('valid_from'),
        data.get('valid_until'),
        id
    ))
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'Kupon güncellendi'})

@app.route('/api/coupons/<id>', methods=['DELETE'])
def delete_coupon(id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE coupons SET is_active = 0 WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'Kupon silindi'})

@app.route('/api/coupons/validate', methods=['POST'])
def validate_coupon():
    data = request.json
    code = data.get('code')
    cart_total = data.get('cart_total', 0)
    now = datetime.now()
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT * FROM coupons 
        WHERE code = ? AND is_active = 1 
        AND (valid_from IS NULL OR valid_from <= ?)
        AND (valid_until IS NULL OR valid_until >= ?)
        AND (usage_limit IS NULL OR used_count < usage_limit)
        AND (minimum_purchase IS NULL OR minimum_purchase <= ?)
    ''', (code, now, now, cart_total))
    coupon = cursor.fetchone()
    conn.close()
    
    if not coupon:
        return jsonify({'valid': False, 'error': 'Geçersiz veya süresi dolmuş kupon'}), 400
    
    coupon_dict = dict(coupon)
    
    # Calculate discount
    if coupon_dict['discount_type'] == 'percentage':
        discount = cart_total * (coupon_dict['discount_value'] / 100)
        if coupon_dict['max_discount']:
            discount = min(discount, coupon_dict['max_discount'])
    else:
        discount = coupon_dict['discount_value']
    
    return jsonify({
        'valid': True,
        'coupon': coupon_dict,
        'discount_amount': discount
    })

# Finance routes
@app.route('/api/finance/transactions', methods=['GET'])
def get_finance_transactions():
    transaction_type = request.args.get('transaction_type')
    category = request.args.get('category')
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 20))
    offset = (page - 1) * limit
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = 'SELECT * FROM finance_transactions WHERE 1=1'
    params = []
    
    if transaction_type:
        query += ' AND transaction_type = ?'
        params.append(transaction_type)
    
    if category:
        query += ' AND category = ?'
        params.append(category)
    
    query += ' ORDER BY transaction_date DESC LIMIT ? OFFSET ?'
    params.extend([limit, offset])
    
    cursor.execute(query, params)
    transactions = cursor.fetchall()
    
    # Get total count
    count_query = 'SELECT COUNT(*) as total FROM finance_transactions WHERE 1=1'
    count_params = []
    
    if transaction_type:
        count_query += ' AND transaction_type = ?'
        count_params.append(transaction_type)
    
    if category:
        count_query += ' AND category = ?'
        count_params.append(category)
    
    cursor.execute(count_query, count_params)
    total = cursor.fetchone()['total']
    
    conn.close()
    
    return jsonify({
        'transactions': [dict(t) for t in transactions],
        'total': total,
        'page': page,
        'limit': limit,
        'totalPages': (total + limit - 1) // limit
    })

@app.route('/api/finance/transactions', methods=['POST'])
def create_finance_transaction():
    data = request.json
    transaction_type = data.get('transaction_type')
    amount = data.get('amount')
    
    if not transaction_type or amount is None:
        return jsonify({'error': 'Gerekli alanlar eksik'}), 400
    
    transaction_id = str(uuid.uuid4())
    user_id = get_current_user_id()
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO finance_transactions (id, transaction_type, amount, category, description, reference_id, reference_type, payment_method, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        transaction_id,
        transaction_type,
        amount,
        data.get('category'),
        data.get('description'),
        data.get('reference_id'),
        data.get('reference_type'),
        data.get('payment_method'),
        user_id
    ))
    conn.commit()
    conn.close()
    
    return jsonify({
        'id': transaction_id,
        'transaction_type': transaction_type,
        'amount': amount,
        'message': 'Finans işlemi kaydedildi'
    }), 201

@app.route('/api/finance/transactions/<id>', methods=['DELETE'])
def delete_finance_transaction(id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM finance_transactions WHERE id = ?', (id,))
    if cursor.rowcount == 0:
        conn.close()
        return jsonify({'error': 'İşlem bulunamadı'}), 404
    conn.commit()
    conn.close()
    return jsonify({'message': 'İşlem silindi'})

@app.route('/api/finance/summary', methods=['GET'])
def get_finance_summary():
    conn = get_db()
    cursor = conn.cursor()
    
    summary = {}
    
    # Income
    cursor.execute("SELECT SUM(amount) as total FROM finance_transactions WHERE transaction_type = 'income' AND status = 'completed'")
    result = cursor.fetchone()
    summary['total_income'] = result['total'] or 0
    
    # Expenses
    cursor.execute("SELECT SUM(amount) as total FROM finance_transactions WHERE transaction_type = 'expense' AND status = 'completed'")
    result = cursor.fetchone()
    summary['total_expenses'] = result['total'] or 0
    
    # Balance
    summary['balance'] = summary['total_income'] - summary['total_expenses']
    
    # This month income
    now = datetime.now()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).strftime('%Y-%m-%d %H:%M:%S')
    
    cursor.execute('''
        SELECT SUM(amount) as total FROM finance_transactions 
        WHERE transaction_type = 'income' AND status = 'completed'
        AND transaction_date >= ?
    ''', (start_of_month,))
    result = cursor.fetchone()
    summary['month_income'] = result['total'] or 0
    
    # This month expenses
    cursor.execute('''
        SELECT SUM(amount) as total FROM finance_transactions 
        WHERE transaction_type = 'expense' AND status = 'completed'
        AND transaction_date >= ?
    ''', (start_of_month,))
    result = cursor.fetchone()
    summary['month_expenses'] = result['total'] or 0
    
    conn.close()
    
    return jsonify(summary)

# Activity logs routes
@app.route('/api/activity-logs', methods=['GET'])
def get_activity_logs():
    entity_type = request.args.get('entity_type')
    entity_id = request.args.get('entity_id')
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 50))
    offset = (page - 1) * limit
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = '''
        SELECT al.*, u.username as user_name
        FROM activity_logs al
        LEFT JOIN users u ON al.user_id = u.id
        WHERE 1=1
    '''
    params = []
    
    if entity_type:
        query += ' AND al.entity_type = ?'
        params.append(entity_type)
    
    if entity_id:
        query += ' AND al.entity_id = ?'
        params.append(entity_id)
    
    query += ' ORDER BY al.created_at DESC LIMIT ? OFFSET ?'
    params.extend([limit, offset])
    
    cursor.execute(query, params)
    logs = cursor.fetchall()
    
    # Get total count
    count_query = 'SELECT COUNT(*) as total FROM activity_logs WHERE 1=1'
    count_params = []
    
    if entity_type:
        count_query += ' AND entity_type = ?'
        count_params.append(entity_type)
    
    if entity_id:
        count_query += ' AND entity_id = ?'
        count_params.append(entity_id)
    
    cursor.execute(count_query, count_params)
    total = cursor.fetchone()['total']
    
    conn.close()
    
    return jsonify({
        'logs': [dict(l) for l in logs],
        'total': total,
        'page': page,
        'limit': limit,
        'totalPages': (total + limit - 1) // limit
    })

# Product Variations routes
@app.route('/api/products/<product_id>/variations', methods=['GET'])
def get_product_variations(product_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT pv.*, p.name as product_name 
        FROM product_variations pv
        LEFT JOIN products p ON pv.product_id = p.id
        WHERE pv.product_id = ?
        ORDER BY pv.phone_model, pv.color, pv.protection_type
    ''', (product_id,))
    variations = cursor.fetchall()
    conn.close()
    
    return jsonify({'variations': [dict(v) for v in variations]})

@app.route('/api/variations', methods=['POST'])
def create_variation():
    data = request.json
    variation_id = str(uuid.uuid4())
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO product_variations (id, product_id, phone_model, color, protection_type, price, stock_quantity, sku, barcode, is_active)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        variation_id,
        data.get('product_id'),
        data.get('phone_model'),
        data.get('color'),
        data.get('protection_type'),
        data.get('price'),
        data.get('stock_quantity', 0),
        data.get('sku'),
        data.get('barcode'),
        data.get('is_active', True)
    ))
    
    # Log activity
    log_activity(
        'create',
        'variation',
        variation_id,
        f'Created variation for product {data.get("product_id")}'
    )
    
    conn.commit()
    conn.close()
    
    return jsonify({'id': variation_id, 'message': 'Varyasyon oluşturuldu'}), 201

@app.route('/api/variations/<variation_id>', methods=['PUT'])
def update_variation(variation_id):
    data = request.json
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE product_variations 
        SET phone_model = ?, color = ?, protection_type = ?, price = ?, stock_quantity = ?, 
            sku = ?, barcode = ?, is_active = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    ''', (
        data.get('phone_model'),
        data.get('color'),
        data.get('protection_type'),
        data.get('price'),
        data.get('stock_quantity'),
        data.get('sku'),
        data.get('barcode'),
        data.get('is_active'),
        variation_id
    ))
    
    # Log activity
    log_activity(
        'update',
        'variation',
        variation_id,
        f'Updated variation'
    )
    
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'Varyasyon güncellendi'})

@app.route('/api/variations/<variation_id>', methods=['DELETE'])
def delete_variation(variation_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM product_variations WHERE id = ?', (variation_id,))
    
    # Log activity
    log_activity(
        'delete',
        'variation',
        variation_id,
        f'Deleted variation'
    )
    
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'Varyasyon silindi'})

# Export route
@app.route('/api/export/orders', methods=['GET'])
def export_orders():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    format_type = request.args.get('format', 'csv')
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = '''
        SELECT o.order_number, c.name as customer_name, o.status, o.subtotal, o.tax, 
        o.shipping_cost, o.discount, o.total, o.created_at
        FROM orders o
        LEFT JOIN customers c ON o.customer_id = c.id
        WHERE 1=1
    '''
    params = []
    
    if start_date:
        query += ' AND o.created_at >= ?'
        params.append(start_date)
    
    if end_date:
        query += ' AND o.created_at <= ?'
        params.append(end_date)
    
    query += ' ORDER BY o.created_at DESC'
    
    cursor.execute(query, params)
    orders = cursor.fetchall()
    conn.close()
    
    if format_type == 'excel':
        # Convert to Excel
        wb = Workbook()
        ws = wb.active
        ws.title = 'Siparişler'
        ws.append(['Sipariş No', 'Müşteri', 'Durum', 'Ara Toplam', 'Vergi', 'Kargo', 'İndirim', 'Toplam', 'Tarih'])
        for order in orders:
            ws.append([
                order['order_number'],
                order['customer_name'] or '',
                order['status'],
                order['subtotal'],
                order['tax'],
                order['shipping_cost'],
                order['discount'],
                order['total'],
                order['created_at']
            ])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = app.response_class(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=siparisler.xlsx'}
        )
        
        return response
    else:
        # Convert to CSV
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Sipariş No', 'Müşteri', 'Durum', 'Ara Toplam', 'Vergi', 'Kargo', 'İndirim', 'Toplam', 'Tarih'])
        
        for order in orders:
            writer.writerow([
                order['order_number'],
                order['customer_name'] or '',
                order['status'],
                order['subtotal'],
                order['tax'],
                order['shipping_cost'],
                order['discount'],
                order['total'],
                order['created_at']
            ])
        
        response = app.response_class(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=siparisler.csv'}
        )
        
        return response

@app.route('/api/export/products', methods=['GET'])
def export_products():
    format_type = request.args.get('format', 'csv')
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = '''
        SELECT p.name, p.description, p.price, p.cost_price, p.stock_quantity, 
        p.sku, p.barcode, c.name as category_name, p.is_active, p.created_at
        FROM products p
        LEFT JOIN categories c ON p.category_id = c.id
    '''
    
    cursor.execute(query)
    products = cursor.fetchall()
    conn.close()
    
    if format_type == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Ürünler'
        ws.append(['Ürün Adı', 'Açıklama', 'Satış Fiyatı', 'Maliyet', 'Stok', 'SKU', 'Barkod', 'Kategori', 'Durum', 'Oluşturma Tarihi'])
        
        for product in products:
            ws.append([
                product['name'],
                product['description'] or '',
                product['price'],
                product['cost_price'] or '',
                product['stock_quantity'],
                product['sku'] or '',
                product['barcode'] or '',
                product['category_name'] or '',
                'Aktif' if product['is_active'] else 'Pasif',
                product['created_at']
            ])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = app.response_class(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=urunler.xlsx'}
        )
        
        return response
    else:
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Ürün Adı', 'Açıklama', 'Satış Fiyatı', 'Maliyet', 'Stok', 'SKU', 'Barkod', 'Kategori', 'Durum', 'Oluşturma Tarihi'])
        
        for product in products:
            writer.writerow([
                product['name'],
                product['description'] or '',
                product['price'],
                product['cost_price'] or '',
                product['stock_quantity'],
                product['sku'] or '',
                product['barcode'] or '',
                product['category_name'] or '',
                'Aktif' if product['is_active'] else 'Pasif',
                product['created_at']
            ])
        
        response = app.response_class(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=urunler.csv'}
        )
        
        return response

@app.route('/api/export/customers', methods=['GET'])
def export_customers():
    format_type = request.args.get('format', 'csv')
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = 'SELECT * FROM customers ORDER BY name'
    cursor.execute(query)
    customers = cursor.fetchall()
    conn.close()
    
    if format_type == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Müşteriler'
        ws.append(['ID', 'Ad Soyad', 'E-posta', 'Telefon', 'Adres', 'Şehir', 'Vergi No', 'Vergi Dairesi', 'Oluşturma Tarihi', 'Güncelleme Tarihi'])
        
        for customer in customers:
            ws.append([
                customer['id'],
                customer['name'],
                customer['email'] or '',
                customer['phone'] or '',
                customer['address'] or '',
                customer['city'] or '',
                customer['tax_number'] or '',
                customer['tax_office'] or '',
                customer['created_at'],
                customer['updated_at']
            ])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = app.response_class(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=musteriler.xlsx'}
        )
        
        return response
    else:
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Ad Soyad', 'E-posta', 'Telefon', 'Adres', 'Şehir', 'Vergi No', 'Vergi Dairesi'])
        
        for customer in customers:
            writer.writerow([
                customer['name'],
                customer['email'] or '',
                customer['phone'] or '',
                customer['address'] or '',
                customer['city'] or '',
                customer['tax_number'] or '',
                customer['tax_office'] or ''
            ])
        
        response = app.response_class(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=musteriler.csv'}
        )
        
        return response

@app.route('/api/export/finance', methods=['GET'])
def export_finance():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    format_type = request.args.get('format', 'csv')
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = 'SELECT * FROM finance_transactions WHERE 1=1'
    params = []
    
    if start_date:
        query += ' AND transaction_date >= ?'
        params.append(start_date)
    
    if end_date:
        query += ' AND transaction_date <= ?'
        params.append(end_date)
    
    query += ' ORDER BY transaction_date DESC'
    
    cursor.execute(query, params)
    transactions = cursor.fetchall()
    conn.close()
    
    if format_type == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Finans'
        ws.append(['ID', 'Tür', 'Tutar', 'Tarih', 'Kategori', 'Açıklama', 'Referans ID', 'Referans Tür', 'Ödeme Yöntemi', 'Durum', 'Oluşturma Tarihi'])
        
        for t in transactions:
            ws.append([
                t['id'],
                t['transaction_type'],
                t['amount'],
                t['transaction_date'],
                t['category'] or '',
                t['description'] or '',
                t['reference_id'] or '',
                t['reference_type'] or '',
                t['payment_method'] or '',
                t['status'],
                t['created_at']
            ])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = app.response_class(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=finans.xlsx'}
        )
        
        return response
    else:
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Tür', 'Tutar', 'Tarih', 'Kategori', 'Açıklama', 'Ödeme Yöntemi', 'Durum'])
        
        for transaction in transactions:
            writer.writerow([
                transaction['transaction_type'],
                transaction['amount'],
                transaction['transaction_date'],
                transaction['category'] or '',
                transaction['description'] or '',
                transaction['payment_method'] or '',
                transaction['status']
            ])
        
        response = app.response_class(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=finans.csv'}
        )
        
        return response

# PDF Report routes
@app.route('/api/pdf/invoice/<order_id>', methods=['GET'])
def generate_invoice_pdf(order_id):
    STATUS_LABELS = {
        'pending': 'Bekliyor',
        'processing': 'İşleniyor',
        'shipped': 'Kargolandı',
        'delivered': 'Teslim Edildi',
        'cancelled': 'İptal Edildi'
    }

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT o.*, c.name as customer_name, c.email as customer_email, c.phone as customer_phone,
               c.address as customer_address, c.city as customer_city
        FROM orders o
        LEFT JOIN customers c ON o.customer_id = c.id
        WHERE o.id = ?
    ''', (order_id,))
    order = cursor.fetchone()

    if not order:
        conn.close()
        return jsonify({'error': 'Sipariş bulunamadı'}), 404

    cursor.execute('''
        SELECT oi.*, p.name as product_name
        FROM order_items oi
        LEFT JOIN products p ON oi.product_id = p.id
        WHERE oi.order_id = ?
    ''', (order_id,))
    items = cursor.fetchall()

    cursor.execute("SELECT * FROM settings WHERE key = 'company_name'")
    company_name_row = cursor.fetchone()
    cursor.execute("SELECT * FROM settings WHERE key = 'address'")
    company_address_row = cursor.fetchone()
    cursor.execute("SELECT * FROM settings WHERE key = 'phone'")
    company_phone_row = cursor.fetchone()

    conn.close()

    created_at = order['created_at']
    if isinstance(created_at, datetime):
        created_at_str = created_at.strftime('%d.%m.%Y %H:%M')
    else:
        created_at_str = str(created_at)

    status_label = STATUS_LABELS.get(order['status'], order['status'])

    company_name = escape((company_name_row['value'] if company_name_row else 'Şirket') or 'Şirket')
    company_address = escape((company_address_row['value'] if company_address_row else '') or '')
    company_phone = escape((company_phone_row['value'] if company_phone_row else '') or '')

    customer_name = escape(order['customer_name'] or '-')
    customer_phone = escape(order['customer_phone'] or '-')
    customer_address = escape(f"{order['customer_address'] or ''} {order['customer_city'] or ''}".strip() or '-')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
    styles = getSampleStyleSheet()
    story = []

    normal_style = ParagraphStyle('DejaVuNormal', parent=styles['Normal'], fontName='DejaVuSans', fontSize=10, leading=14)
    bold_style = ParagraphStyle('DejaVuBold', parent=styles['Normal'], fontName='DejaVuSans-Bold', fontSize=10, leading=14)
    title_style = ParagraphStyle('InvoiceTitle', parent=styles['Heading1'], fontName='DejaVuSans-Bold', fontSize=26, textColor=colors.HexColor('#1e40af'), alignment=2, spaceAfter=12)
    section_style = ParagraphStyle('SectionHeader', parent=styles['Normal'], fontName='DejaVuSans-Bold', fontSize=11, textColor=colors.white, backColor=colors.HexColor('#1e40af'), leading=16, leftIndent=6, spaceBefore=6, spaceAfter=6)
    small_style = ParagraphStyle('Small', parent=styles['Normal'], fontName='DejaVuSans', fontSize=9, leading=12, textColor=colors.HexColor('#666666'))

    # Header
    story.append(Paragraph('FATURA', title_style))
    story.append(Paragraph(
        f'<font name="DejaVuSans-Bold">Fatura No:</font> {escape(order["order_number"])}<br/>'
        f'<font name="DejaVuSans-Bold">Tarih:</font> {created_at_str}<br/>'
        f'<font name="DejaVuSans-Bold">Durum:</font> {status_label}',
        normal_style
    ))
    story.append(Spacer(1, 0.2 * inch))

    # Company & customer info side by side
    company_text = f'<font name="DejaVuSans-Bold">{company_name}</font><br/>{company_address}<br/>Tel: {company_phone}'
    customer_text = f'<font name="DejaVuSans-Bold">Fatura Bilgileri</font><br/>{customer_name}<br/>{customer_phone}<br/>{customer_address}'

    info_table = Table([[Paragraph(company_text, normal_style), Paragraph(customer_text, normal_style)]], colWidths=[3.3*inch, 3.3*inch])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#f3f4f6')),
        ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#eff6ff')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.25 * inch))

    # Items
    story.append(Paragraph('Sipariş Kalemleri', section_style))
    story.append(Spacer(1, 0.05 * inch))

    table_data = [[Paragraph('Ürün', bold_style), Paragraph('Miktar', bold_style), Paragraph('Birim Fiyat', bold_style), Paragraph('Toplam', bold_style)]]
    for item in items:
        table_data.append([
            Paragraph(escape(str(item['product_name'] or '-')), normal_style),
            Paragraph(str(item['quantity']), normal_style),
            Paragraph(f"₺{item['unit_price']:.2f}", normal_style),
            Paragraph(f"₺{item['total_price']:.2f}", normal_style),
        ])

    totals_rows = [
        ['', '', Paragraph('Ara Toplam', bold_style), Paragraph(f"₺{order['subtotal']:.2f}", bold_style)],
        ['', '', Paragraph('Vergi', bold_style), Paragraph(f"₺{order['tax']:.2f}", bold_style)],
        ['', '', Paragraph('Kargo', bold_style), Paragraph(f"₺{order['shipping_cost']:.2f}", bold_style)],
        ['', '', Paragraph('İndirim', bold_style), Paragraph(f"₺{order['discount']:.2f}", bold_style)],
        ['', '', Paragraph('GENEL TOPLAM', bold_style), Paragraph(f"₺{order['total']:.2f}", bold_style)],
    ]
    for row in totals_rows:
        table_data.append(row)

    n_items = len(items)
    header_row = 0
    first_item_row = 1
    first_total_row = n_items + 1

    col_widths = [3.2*inch, 0.9*inch, 1.2*inch, 1.2*inch]
    items_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, header_row), (-1, header_row), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, header_row), (-1, header_row), colors.white),
        ('FONTNAME', (0, header_row), (-1, header_row), 'DejaVuSans-Bold'),
        ('FONTNAME', (0, first_item_row), (-1, first_total_row - 1), 'DejaVuSans'),
        ('ALIGN', (1, first_item_row), (1, -1), 'CENTER'),
        ('ALIGN', (2, first_item_row), (3, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, header_row), (-1, header_row), 10),
        ('TOPPADDING', (0, first_item_row), (-1, -1), 8),
        ('BOTTOMPADDING', (0, first_item_row), (-1, -1), 8),
        ('GRID', (0, header_row), (-1, first_total_row - 1), 0.5, colors.HexColor('#d1d5db')),
        ('BOX', (0, header_row), (-1, first_total_row - 1), 1, colors.HexColor('#1e40af')),
        ('LINEABOVE', (2, first_total_row), (3, first_total_row), 1, colors.HexColor('#9ca3af')),
        ('BACKGROUND', (0, first_total_row), (-1, -1), colors.HexColor('#f3f4f6')),
        ('FONTNAME', (2, first_total_row), (3, -1), 'DejaVuSans-Bold'),
    ]))
    story.append(items_table)
    story.append(Spacer(1, 0.2 * inch))

    if order['notes']:
        story.append(Paragraph('Sipariş Notları', section_style))
        story.append(Spacer(1, 0.05 * inch))
        story.append(Paragraph(escape(str(order['notes'])), normal_style))
        story.append(Spacer(1, 0.15 * inch))

    story.append(Paragraph('Teşekkür ederiz.', small_style))

    doc.build(story)
    buffer.seek(0)

    response = app.response_class(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=fatura_{order["order_number"]}.pdf'}
    )

    return response

@app.route('/api/pdf/stock-report', methods=['GET'])
def generate_stock_report_pdf():
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT p.name, p.sku, p.stock_quantity, p.price, c.name as category_name
        FROM products p
        LEFT JOIN categories c ON p.category_id = c.id
        ORDER BY p.name
    ''')
    products = cursor.fetchall()
    conn.close()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    story = []

    normal_style = ParagraphStyle('StockNormal', parent=styles['Normal'], fontName='DejaVuSans', fontSize=10, leading=13)
    bold_style = ParagraphStyle('StockBold', parent=styles['Normal'], fontName='DejaVuSans-Bold', fontSize=10, leading=13)
    title_style = ParagraphStyle('StockTitle', parent=styles['Heading1'], fontName='DejaVuSans-Bold', fontSize=22, textColor=colors.HexColor('#1e40af'), alignment=1, spaceAfter=16)

    story.append(Paragraph('STOK RAPORU', title_style))
    story.append(Paragraph(f'Oluşturma Tarihi: {datetime.now().strftime("%d.%m.%Y %H:%M")}', normal_style))
    story.append(Spacer(1, 0.25 * inch))

    table_data = [[Paragraph('Ürün', bold_style), Paragraph('SKU', bold_style), Paragraph('Stok', bold_style), Paragraph('Fiyat', bold_style), Paragraph('Kategori', bold_style)]]
    for product in products:
        table_data.append([
            Paragraph(escape(str(product['name'])), normal_style),
            Paragraph(escape(str(product['sku'] or '-')), normal_style),
            Paragraph(str(product['stock_quantity']), normal_style),
            Paragraph(f"₺{product['price']:.2f}", normal_style),
            Paragraph(escape(str(product['category_name'] or '-')), normal_style),
        ])

    table = Table(table_data, colWidths=[3*inch, 1.2*inch, 0.8*inch, 1*inch, 1.5*inch], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1e40af')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
    ]))
    story.append(table)

    doc.build(story)
    buffer.seek(0)

    response = app.response_class(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': 'attachment; filename=stok_raporu.pdf'}
    )

    return response

# Backup / Restore / Trash routes
BACKUP_TABLES = [
    'categories', 'products', 'customers', 'suppliers', 'orders', 'order_items',
    'returns', 'return_items', 'coupons', 'finance_transactions',
    'product_variations', 'stock_movements', 'settings'
]

def _backup_to_dict(conn):
    cursor = conn.cursor()
    backup = {}
    for table in BACKUP_TABLES:
        cursor.execute(f'SELECT * FROM {table}')
        backup[table] = [_serialize_row(r) for r in cursor.fetchall()]
    cursor.execute('SELECT id, username, email, full_name, role, theme, language, created_at, updated_at FROM users')
    backup['users'] = [_serialize_row(r) for r in cursor.fetchall()]
    return backup

def _write_local_backup(backup):
    filename = f'backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
    path = os.path.join(BACKUP_DIR, filename)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(backup, f, ensure_ascii=False, indent=2, default=str)
    return filename

@app.route('/api/backup', methods=['GET'])
@admin_required
def export_backup():
    conn = get_db()
    backup = _backup_to_dict(conn)
    conn.close()
    backup['exported_at'] = datetime.now().isoformat()

    if request.args.get('download') == '1':
        output = io.BytesIO(json.dumps(backup, ensure_ascii=False, indent=2, default=str).encode('utf-8'))
        return app.response_class(
            output.getvalue(),
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment; filename=backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'}
        )

    return jsonify(backup)

@app.route('/api/backup', methods=['POST'])
@admin_required
def create_local_backup():
    conn = get_db()
    backup = _backup_to_dict(conn)
    conn.close()
    backup['exported_at'] = datetime.now().isoformat()
    filename = _write_local_backup(backup)
    log_activity('backup', 'system', '', {'filename': filename})
    return jsonify({'message': 'Yedek oluşturuldu', 'filename': filename})

@app.route('/api/backup/restore', methods=['POST'])
@admin_required
def restore_backup():
    data = request.json or {}
    tables = data.get('tables')
    if not isinstance(tables, dict):
        return jsonify({'error': 'Geri yüklenecek tablolar gerekli'}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        for table, rows in tables.items():
            if not rows:
                continue
            if table == 'settings':
                for row in rows:
                    cursor.execute('''
                        INSERT INTO settings (key, value, updated_at)
                        VALUES (?, ?, CURRENT_TIMESTAMP)
                        ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = CURRENT_TIMESTAMP
                    ''', (row.get('key'), row.get('value')))
            elif table == 'users':
                continue  # Do not overwrite users or passwords
            else:
                _upsert_table(cursor, table, rows)
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': f'Geri yükleme hatası: {str(e)}'}), 400
    conn.close()

    log_activity('restore', 'backup', '', {'tables': list(tables.keys())})
    return jsonify({'message': 'Yedek başarıyla geri yüklendi'})

@app.route('/api/backup/list', methods=['GET'])
@admin_required
def list_backups():
    files = [f for f in os.listdir(BACKUP_DIR) if f.endswith('.json')]
    files.sort(reverse=True)
    return jsonify(files)

@app.route('/api/trash', methods=['GET'])
@admin_required
def get_trash():
    conn = get_db()
    cursor = conn.cursor()
    result = {}
    for table, name_col in [('products', 'name'), ('customers', 'name'), ('categories', 'name'), ('orders', 'order_number')]:
        cursor.execute(f'SELECT id, {name_col} as name, deleted_at FROM {table} WHERE deleted_at IS NOT NULL ORDER BY deleted_at DESC')
        result[table] = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify(result)

@app.route('/api/trash/<entity>/<id>/restore', methods=['POST'])
@admin_required
def restore_trash_item(entity, id):
    if entity not in ('products', 'customers', 'categories', 'orders'):
        return jsonify({'error': 'Geçersiz varlık tipi'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(f'UPDATE {entity} SET deleted_at = NULL, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (id,))
    if cursor.rowcount == 0:
        conn.close()
        return jsonify({'error': 'Öğe bulunamadı'}), 404
    conn.commit()
    conn.close()

    log_activity('restore', entity, id, {})
    return jsonify({'message': 'Öğe çöp kutusundan geri yüklendi'})

# Initialize database at startup
init_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=True, host='0.0.0.0', port=port)
